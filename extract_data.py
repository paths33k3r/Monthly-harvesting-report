import openpyxl
import json

def extract_grouped_data(file_path):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True) # Only need values for this tailored extraction
    
    sheet_name = wb.sheetnames[0] # Target first sheet 'OVERALL BY GANG COMPARISON YTD2'
    print(f"Extracting structured data from: {sheet_name}")
    ws = wb[sheet_name]
    
    # Let's map exactly based on the debug output for this sheet:
    # Block is in Column A (Index 1)
    # Phase/OP is in Column B (Index 2)
    # Total HA is in Column C (Index 3)
    
    block_col = 1
    op_col = 2
    ha_col = 3
    
    # Find where the actual data starts (look for the first OP year row)
    header_row_idx = 1
    for r in range(1, 20):
        val = ws.cell(row=r, column=op_col).value
        # Look for something like "2010" in the O/P column
        if val and (isinstance(val, int) or (isinstance(val, str) and val.strip().isdigit())):
            header_row_idx = r - 1
            break
            
    print(f"Columns determined: Block={block_col}, O/P={op_col}, HA={ha_col}")
    print(f"Starting data scan at row {header_row_idx + 1}")

    # 2. Extract Data Hierarchically
    grouped_data = [] 
    current_group = None
    grand_total = 0.0
    
    for r in range(header_row_idx + 1, ws.max_row + 1):
        block_val = ws.cell(row=r, column=block_col).value
        op_val = ws.cell(row=r, column=op_col).value
        ha_val = ws.cell(row=r, column=ha_col).value
        
        # Determine if this row is an O/P Group Header (has O/P year, but no block ID)
        if op_val and (isinstance(op_val, int) or (isinstance(op_val, str) and op_val.strip().isdigit())) and not block_val:
            year_str = str(op_val).strip()
            
            sub_val = 0.0
            if ha_val and isinstance(ha_val, (int, float)):
                sub_val = float(ha_val)
                
            current_group = {
                "op_year": year_str,
                "initial_subtotal": sub_val, 
                "blocks": []
            }
            grouped_data.append(current_group)
            
        # Determine if this row is a Block Row (has block ID, O/P year, and HA, and we are inside a group)
        elif block_val and op_val and ha_val and current_group is not None:
             b_idx = str(block_val).strip()
             try:
                 h_val = float(ha_val)
                 current_group["blocks"].append({
                     "block_id": b_idx,
                     "ha": h_val,
                 })
             except ValueError:
                 pass # Not a valid float HA constraint
             
        # Determine if this is the Grand Total Row ("All Blocks" in column A)
        elif block_val and "ALL BLOCKS" in str(block_val).upper():
             val_c = ws.cell(row=r, column=2).value # In the screenshot, the total might be in col 2 or 3 for the grand total row. Let's check both
             if isinstance(val_c, (int, float)): grand_total = float(val_c)
             elif isinstance(ha_val, (int, float)): grand_total = float(ha_val)
             
             print(f"Found Grand Total: {grand_total}")
             break # End of table
             
    output_dict = {
        "groups": grouped_data,
        "initial_grand_total": grand_total
    }
    
    # Debug print structure
    print(f"Extracted {len(grouped_data)} O/P Groups.")
    for grp in grouped_data:
         print(f" - O/P {grp['op_year']}: {len(grp['blocks'])} blocks")
         
    output_file = "grouped_data.json"
    print(f"Writing structured data to {output_file}...")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output_dict, f, indent=2)
        
    print("Extraction successful.")

if __name__ == "__main__":
    # From the earlier conversation, we know the main file has a typo ('Havesting')
    extract_grouped_data("Havesting Performance Jan 2026.xlsx")
