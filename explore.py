import openpyxl

excel_file = "Havesting Performance Jan 2026.xlsx"
print(f"Loading {excel_file}...")
wb = openpyxl.load_workbook(excel_file, data_only=False)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(f"{cell.coordinate}: {cell.value} (Type: {type(cell.value).__name__})")
