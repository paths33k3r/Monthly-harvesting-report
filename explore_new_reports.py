import openpyxl
import json

excel_file = "Havesting Performance Jan 2026.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)

target_sheets = ["OVERALL BY GANG COMPARISON YTD2", "OVERALL BY GANG COMPARISON V3"]

for sheet_name in target_sheets:
    if sheet_name in wb.sheetnames:
        print(f"\n==========================================")
        print(f"Sheet: {sheet_name}")
        print(f"==========================================")
        ws = wb[sheet_name]
        for r in range(1, 15):
            row_vals = []
            for c in range(1, 30):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                row_vals.append(str(val) if val is not None else "")
            print(f"Row {r:02d}:", " | ".join(row_vals))
