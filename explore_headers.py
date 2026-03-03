import openpyxl

excel_file = "Havesting Performance Jan 2026.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=False)
wb_data = openpyxl.load_workbook(excel_file, data_only=True)

sheet1 = wb.sheetnames[0]
print(f"Sheet: {sheet1}")

ws = wb[sheet1]
ws_data = wb_data[sheet1]

for r in range(1, 20):
    row_vals = []
    for c in range(1, 40):
        cell = ws.cell(row=r, column=c)
        d_cell = ws_data.cell(row=r, column=c)
        val = cell.value
        data_val = d_cell.value
        if val is not None:
            if isinstance(val, str) and val.startswith('='):
                row_vals.append(f"FORMULA({val})[val={data_val}]")
            else:
                row_vals.append(str(val))
        else:
            row_vals.append("")
    print(f"Row {r}:", " | ".join(row_vals))
