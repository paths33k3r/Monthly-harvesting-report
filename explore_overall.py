import openpyxl

file_path = "Havesting Performance Jan 2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet_name = wb.sheetnames[0]
print(f"--- RAW ROW DUMP FOR {sheet_name} ---")
ws = wb[sheet_name]
for r in range(1, 45):
    row_vals = []
    for c in range(1, 15): # First 15 columns
        val = ws.cell(row=r, column=c).value
        # Shorten none to empty string
        row_vals.append(str(val) if val is not None else "")
    print(f"Row {r:02d}: {row_vals}")
