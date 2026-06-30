import openpyxl, json, re, datetime

SRC = r'C:\Users\user\Desktop\New folder\1. Logs Species Summary.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

def norm(s):
    return re.sub(r'[^A-Z0-9]', '', str(s).upper()) if s is not None else ''

def isodate(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = str(v or '').strip()
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    return m.group(0) if m else s

batches = {}   # batchNo -> {date, detailed, subtotals:{ "CAT|GRADE":[qty,vol] }, total:[qty,vol]}

for sn in wb.sheetnames:
    if sn.upper().startswith('YEAR ') or sn.upper().endswith('TEMPLATE'):
        continue
    ws = wb[sn]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # ACMG summary sheet?
    is_acmg = False
    hdr_idx = None
    for i, r in enumerate(rows[:8]):
        joined = norm('|'.join('' if c is None else str(c) for c in r))
        if 'BATCHNO' in joined and 'DELIVER' in joined:
            is_acmg = True; hdr_idx = i; break
        if 'SPECIESCATEGORY' in joined and 'GRADE' in joined:
            hdr_idx = i; break

    if is_acmg:
        for r in rows[hdr_idx+1:]:
            if len(r) < 5: continue
            bno = str(r[2]).strip() if r[2] else ''
            if not bno or not re.match(r'^KU', bno, re.I): continue
            qty = float(r[3] or 0); vol = float(r[4] or 0)
            batches[bno] = {'date': isodate(r[1]), 'detailed': False,
                            'species': 'ACMG',
                            'subtotals': {'ACMG|ACMG': [qty, vol]},
                            'total': [qty, vol], 'sheet': sn}
        continue

    if hdr_idx is None:
        continue
    # KU detailed sheet
    batchNo = sn; date = ''
    for r in rows[:6]:
        for ci, c in enumerate(r):
            cs = str(c) if c is not None else ''
            if 'BATCH' in cs.upper() and ':' in cs:
                m = re.search(r'KU[0-9A-Z]+', cs.upper())
                if m: batchNo = m.group(0)
            if 'DELIVER' in cs.upper():
                # date is usually the next non-empty cell in the row
                for c2 in r[ci+1:]:
                    if c2 not in (None, ''):
                        date = isodate(c2); break
    sub = {}        # CATEGORY|GRADE -> [qty,vol]
    subsp = {}      # SPECIES|GRADE  -> [qty,vol]  (for old species-based invoices)
    cur_cat = ''
    for r in rows[hdr_idx+1:]:
        cells = list(r) + [None]*(5-len(r)) if len(r) < 5 else r
        cat = str(cells[0]).strip() if cells[0] else ''
        sp  = str(cells[1]).strip() if cells[1] else ''
        gr  = str(cells[2]).strip() if cells[2] else ''
        qv  = cells[3]; vv = cells[4]
        line = norm('|'.join('' if c is None else str(c) for c in r))
        if 'SUBTOTAL' in line or 'GRANDTOTAL' in line:
            cur_cat = ''  # reset at group boundary
            continue
        if cat: cur_cat = cat
        usecat = cat or cur_cat
        if not gr: continue
        q = float(qv) if isinstance(qv,(int,float)) else 0.0
        v = float(vv) if isinstance(vv,(int,float)) else 0.0
        if q == 0 and v == 0: continue
        key = usecat + '|' + gr
        if key not in sub: sub[key] = [0.0, 0.0]
        sub[key][0] += q; sub[key][1] += v
        if sp:
            skey = sp + '|' + gr
            if skey not in subsp: subsp[skey] = [0.0, 0.0]
            subsp[skey][0] += q; subsp[skey][1] += v
    tot = [round(sum(x[0] for x in sub.values()),3), round(sum(x[1] for x in sub.values()),3)]
    sub = {k:[round(x[0],3), round(x[1],3)] for k,x in sub.items()}
    subsp = {k:[round(x[0],3), round(x[1],3)] for k,x in subsp.items()}
    batches[batchNo] = {'date': date, 'detailed': True, 'subtotals': sub, 'subspecies': subsp, 'total': tot, 'sheet': sn}

with open('batches.json','w') as f:
    json.dump(batches, f, indent=1)

print('Parsed', len(batches), 'batches')
det = sum(1 for b in batches.values() if b['detailed'])
print('detailed:', det, ' summary:', len(batches)-det)
# period index: how many batches per delivery year-month
from collections import Counter
per = Counter(b['date'][:7] for b in batches.values() if b['date'])
print('\nBatches by delivery month:')
for ym in sorted(per): print('  ', ym, per[ym])
print('\nEarliest batch date:', min((b['date'] for b in batches.values() if b['date']), default='?'))
# show a few
for bno in ['KU1224A01','KU1223A02','KU1223A03']:
    if bno in batches:
        b = batches[bno]
        print('\n', bno, b['date'], 'total', b['total'])
        for k,v in b['subtotals'].items(): print('   ', k, v)
