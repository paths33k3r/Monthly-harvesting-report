import openpyxl

excel_file = "Havesting Performance Jan 2026.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
print("Sheet names:", wb.sheetnames)

ws = wb["OVERALL BY GANG COMPARISON YTD2"]
print("\nFirst column in YTD2 (first 50 rows):")
for r in range(5, 50):
    print(f"Row {r:02d}: Col1={ws.cell(r, 1).value}, Col2={ws.cell(r, 2).value}")

ws2 = wb["OVERALL BY GANG COMPARISON V3"]
print("\nFirst column in V3 (first 50 rows):")
for r in range(5, 50):
    val = ws2.cell(r, 1).value
    if val:
        print(f"Row {r:02d}: Col1={val}")
