import openpyxl

excel_file = "Havesting Performance Jan 2026.xlsx"
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb["OVERALL BY GANG COMPARISON YTD2"]

print("\nYTD2 last 30 rows:")
max_row = ws.max_row
for r in range(max_row - 30, max_row + 1):
    row_vals = []
    for c in range(1, 15):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_vals.append(str(val))
    if len(row_vals) > 0:
        print(f"Row {r:03d}:", " | ".join(row_vals))

